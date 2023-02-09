from collections import Counter
import csv
from datetime import datetime
import xlrd
import os
from models import Depot, Measure, Train, User
from openpyxl import Workbook

from my_engine import session_scope

def parsing_file_for_struct(file) -> dict:
    """Функция парсинга файла для создания структуры поездов <-> маршрутов """
    file_type: str = file.filename.split(".")[-1]
    file_name = file.filename

    if file_type.lower() in ("xls"):
        if not os.path.exists("files_struct"):
            os.mkdir("files_struct")
        file.save(f"files_struct/{file_name}")

        book = xlrd.open_workbook(f"files_struct/{file_name}")
        sh = book.sheet_by_index(0)

        result = {"trains": []} # Данные полученные в результате парсинга файла


        # Узнаём id депо из файла
        depo_name = [_ for _ in sh._cell_values[0] if _][0].replace('"', "").split()[-1].lower()
        with session_scope() as session:
            depo_id = session.query(Depot.id).filter(Depot.name.like(f"%{depo_name}%")).all()
        
        if not depo_id:
            return {"status": 400, f"message": f"Депо '{depo_name}' не найдено. Проверьте файл или сообщите об ошибке"}
        elif len(depo_id) > 1:
            return {"status": 400, f"message": f"Найдено больше 1 депо"}
        else:
            result["depo_id"] = depo_id[0][0]


        # Узнаём дату из файла
        for date in [_ for _ in sh._cell_values[2] if _]:
            try:
                date = datetime.strptime(date, "%d.%m.%Y").strftime("%Y-%m-%d")
            except:
                pass
        
        if date:
            result["date_placements"] = date
        else:
            return {"status": 400, "message": "Дата не обнаружена в файле"}


        # Узнаём номера маршрутов <-> номера вагонов
        for item in [_[:-1] for _ in sh._cell_values[5:]]:
            if "ПАССАЖИРСКОЕ ДВИЖЕНИЕ" in item[0]:
                break

            if [_ for _ in item[1:] if _]:
                wagon_numbers = [{"wagon_number": _, "reverse": False} for _ in item[1:] if _]
                count_wagons = len(wagon_numbers)
                route_number = item[0]

                result["trains"].append({
                    "type_id": 2 if count_wagons <= 5 else 1,
                    "count_wagons": count_wagons, # из файла
                    "route_number": route_number,
                    "wagon_numbers": wagon_numbers
                })

        return {"status": 200, "data": result}
    else:
        return {"status": 400, "message": f'Неверный формат файла: "{file_type}", загружайте "xls"'}


def get_data_from_file(file_data, login: str, date: str, route_number: str) -> dict:
    """Загрузить данные из файла и отправка их в верном формате в таблицу (на фронт)"""
    with session_scope() as session:
        type_id, count_wagons = session.query(Train.type_id, Train.count_wagons).filter(Train.date_placement == date,
                                                                                        Train.route_number == route_number,
                                                                                        Train.depo_id == session.query(User.depo_id).filter(User.ad_login == login).one()[0]).first()

    count_of_measure = count_wagons * 12 if type_id == 2 else count_wagons * 8
    file_type: str = file_data.filename.split(".")[-1]
    file_name: str = file_data.filename
    file_path = f"files_data/{file_name}"
    if file_type.lower() in ("csv"):
        if not os.path.exists("files_data"):
            os.mkdir("files_data")
        file_data.save(file_path)

        try:
            with open(file_path) as f:
                temp_data = ['?' for _ in range(61)]
                value_dict = dict((int(_[0]), int(_[1])) for _ in csv.reader(f, delimiter=';') if _)
                for i, v in value_dict.items():
                    temp_data[i - 1] = v
        except:
            return {"status": 400, "message": "Файл некорректный!"}
        temp_ambient = temp_data.pop(0) # для <=5 вагонов, красная если single_temp > temp_ambient + 35 или single_temp отличается больше чем на 5 от остальных в вагоне
        
        if not isinstance(temp_ambient, int):
            return {"status": 400, "message": "Температура окружающей среды не определена"}
        
        temp_data = temp_data[:count_of_measure // 2] + list(reversed(temp_data[count_of_measure // 2:]))
# для заполнения в любом случае
#         return {"status": 200,
#                 "values": [
#                     {
#                         "html_class_name": "left_axle_box" if type_id == 2 else "bearing80",
#                         # "values": temp_data[:count_of_measure // 2]
#                         "values": [{"single_value": _, "extreme": True if _ > temp_ambient + 35 else False} for _ in temp_data[:count_of_measure // 2]]
#                     },
#                     {
#                         "html_class_name": "right_axle_box" if type_id == 2 else "bearing30",
#                         # "values": temp_data[count_of_measure // 2:]
#                         "values": [{"single_value": _, "extreme": True if _ > temp_ambient + 35 else False} for _ in temp_data[count_of_measure // 2:]]
#                     }
#                 ],
#                 "temp_ambient": temp_ambient,
#                 "message": None if len(temp_data) == count_of_measure else f"Количество замеров в файле некорректно. \
# Количество вагонов = {count_wagons}, замеров должно быть: 1 (Замер рамы тележки) + ({count_wagons} * {'12' if type_id == 2 else '8'}) = \
# {(count_wagons * 12 if type_id == 2 else count_wagons * 8) + 1}, а в файле {len(temp_data) + 1} замеров"}
    
    # не заполнять если не совпадает
        if not len(temp_data) == count_of_measure:
            return {"status": 400, "message": f"Количество замеров в файле некорректно. \
Количество вагонов = {count_wagons}, замеров должно быть: 1 (Замер рамы тележки) + ({count_wagons} * {'12' if type_id == 2 else '8'}) = \
{(count_wagons * 12 if type_id == 2 else count_wagons * 8) + 1}, а в файле {len(temp_data) + 1} замеров. Перепроверьте данные"}
        
        return {"status": 200,
                "values": [
                    {
                        "html_class_name": "left_axle_box" if type_id == 2 else "bearing80",
                        "values": [{"single_value": _, "extreme": (True if _ > temp_ambient + 35 else False) if isinstance(_, int) else False} for _ in temp_data[:count_of_measure // 2]]
                    },
                    {
                        "html_class_name": "right_axle_box" if type_id == 2 else "bearing30",
                        "values": [{"single_value": _, "extreme": (True if _ > temp_ambient + 35 else False) if isinstance(_, int) else False} for _ in temp_data[count_of_measure // 2:]]
                    }
                ],
                "temp_ambient": temp_ambient,
                "message": None if temp_data.count("?") == 0 else f"Количество пропусков = {temp_data.count('?')}. Обратите внимание на знаки вопроса"}
    
    else:
        return {"status": 400, "message": f'Неверный формат файла: "{file_type}", загружайте "csv"'}


def get_count_measurement() -> dict():
    """Получить данные о количестве замеров всех вагонов в период с 11.09.2022 по сегодняшний день"""
    wb = Workbook()
    # grab the active worksheet
    ws1 = wb.create_sheet("list1", 0)
    ws2 = wb.create_sheet("list2", 1)

    with session_scope() as session:
        data = session.query(Measure.id, Measure.wagon_numbers, Measure.data, Measure.datetime_measure, Train.route_number).\
            join(Train, Train.id == Measure.train_id).\
                filter(Measure.datetime_measure > '2022-09-12').all()
    
    list_ = [{"id": _[0], 
              "wagon_numbers": _[1]["wagon_numbers"], 
              "data1": _[2]["5"],
              "data1": _[2]["6"],
              "route_number": _[4],
              "date": _[3].strftime("%Y-%m-%d")} for _ in data]
    
    list_measurement_wagons = []
    for item in list_:
        for wagon in item["wagon_numbers"]:
            list_measurement_wagons.append({"wagon_number": wagon["wagon_number"], "route_number": item["route_number"], "date": item["date"]})

    for num, count in Counter([_["wagon_number"] for _ in list_measurement_wagons]).items():
        ws1.append([num, count])

    for item in list_measurement_wagons:
        item["count_measurement"] = Counter([_["wagon_number"] for _ in list_measurement_wagons])[item["wagon_number"]]


    for item in list_measurement_wagons:
        ws2.append([item["wagon_number"], item["count_measurement"], item["date"], item["route_number"]])

    if not os.path.exists("../analitics_files"):
        os.mkdir("../analitics_files")
    
    path = "../analitics_files/выгрузка.xlsx"
    wb.save(path)

    return {"status": 200, "path": "../analitics_files/выгрузка.xlsx"}
